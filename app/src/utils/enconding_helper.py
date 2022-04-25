import openpyxl
import chardet


# Dead code, for illustration 
# try to find the set of encondings
# used once to read the excel file and returns a list of encodings found.
# @returns ['ISO-8859-1' 'Windows-1252', 'ISO-8859-9', 'Windows-1254']
def get_distinct_enconding_of_each_shell(target_data_path: str) -> list[str]:
    wrkbk = openpyxl.load_workbook(target_data_path)
    sh = wrkbk.active
    encondings = []
    for i in range(1, sh.max_row + 1):

        for j in range(1, sh.max_column + 1):

            cell_obj = sh.cell(row=i, column=j)
            raw_data = cell_obj.value
            result = chardet.detect(raw_data.encode())
            charenc = result["encoding"]
            if charenc != "ascii" and charenc != "utf-8":
                encondings.append(charenc)
    return encondings
